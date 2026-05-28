import re
import os
import json
import logging
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from typing import Dict, Any, List, Optional

import dspy
from src.langchain_integration import RAGPipeline
from .signatures import TestPlanGenerationSignature, TestCodeGenerationSignature

logger = logging.getLogger(__name__)


class RAGPlaywrightGenerator(dspy.Module):
    def __init__(self, region: str = "ap-northeast-1"):
        super().__init__()

        self.rag_pipeline = RAGPipeline(
            index_name="playwright_code_index",
            region=region
        )

        self.plan_generator = dspy.Predict(TestPlanGenerationSignature)
        self.code_generator = dspy.Predict(TestCodeGenerationSignature)

        self.save_base_path = "/home/ec2-user/AI/output_codes"
        if not os.path.exists(self.save_base_path):
            os.makedirs(self.save_base_path)
            logger.info(f"Created output directory: {self.save_base_path}")

    # ── 인덱싱 ──────────────────────────────────────────────────────────────

    def index_documents(self, documents: List[Any], file_tree: Optional[str] = None):
        """저장소 코드를 RAG 벡터 DB에 등록"""
        try:
            ids = self.rag_pipeline.add_documents(documents)
            self.file_tree = file_tree
            logger.info(f"Successfully indexed {len(ids)} documents.")
        except Exception as e:
            logger.error(f"Indexing failed: {e}")
            raise

    # ── STEP 1: 계획서 생성 ──────────────────────────────────────────────────

    def generate_plan_only(self, requirement: str, **kwargs) -> Dict[str, Any]:
        """STEP 1: RAG 검색 없이 시나리오명만으로 계획서 생성"""
        execution_id = kwargs.get("execution_id", 0)
        plan_index = kwargs.get("plan_index", 0)
        try:
            # 계획서 생성은 레포 코드 없이 시나리오 도메인 지식만으로 생성
            code_context = "No repository context. Generate test plan based on scenario description only."
            plan_prediction = self.plan_generator(
                requirement=requirement,
                code_context=code_context,
                scenario_serial=kwargs.get("scenario_serial", "00"),
                scenario_attempt=kwargs.get("scenario_attempt", "01"),
            )
            raw = getattr(plan_prediction, "test_plan", "[]")
            logger.info(f"[{execution_id}] test_plan raw type={type(raw)}, value={str(raw)[:300]}")
            test_cases = self._parse_test_plan(raw)
            logger.info(f"[{execution_id}] test_cases count={len(test_cases)}")

            timestamp = datetime.now().strftime("%Y%m%d")
            plan_path = os.path.join(self.save_base_path, f"{timestamp}_{execution_id}_plan_{plan_index+1}.xlsx")
            _save_plan_as_excel(test_cases, plan_path)

            return {"status": "success", "saved_plan": plan_path, "test_cases": test_cases}
        except Exception as e:
            return {"status": "error", "message": str(e)}

    # ── STEP 2: 테스트 코드 생성 ─────────────────────────────────────────────

    # 시나리오별 영어 검색 쿼리 (레포 코드가 영어라 한글 쿼리보다 유사도 높음)
    SCENARIO_SEARCH_QUERIES = {
        "01": ["signup register route path", "signup form input validation", "register API endpoint"],
        "02": ["login route path URL", "login form input placeholder submit button", "auth login API endpoint"],
        "03": ["password reset forgot route", "password reset form input", "password reset API"],
        "04": ["logout route", "logout button handler", "logout API session"],
        "05": ["profile edit form input", "profile update API endpoint", "mypage route"],
        "06": ["password change form input", "password update API endpoint"],
        "07": ["role permission access control route", "protected route guard"],
        "08": ["session token expiry refresh", "token validation middleware"],
        "09": ["post create form input", "post write API endpoint", "board route"],
        "10": ["post edit update delete form", "post update delete API endpoint"],
        "11": ["comment create edit delete form", "comment API endpoint"],
        "12": ["like favorite toggle button", "like API endpoint"],
        "13": ["search input form", "search API endpoint filter"],
        "14": ["filter sort dropdown button", "filter API query params"],
        "15": ["responsive layout media query viewport", "mobile header nav"],
        "16": ["browser compatibility CSS font", "header navigation layout"],
        "17": ["404 500 error page not found", "error boundary route"],
        "18": ["network offline error handling", "fetch error catch"],
        "19": ["loading spinner skeleton delay", "async API slow response"],
        "20": ["API error response handling", "fetch catch error message"],
        "21": ["AB test variant feature flag", "experiment toggle"],
        "22": ["form validation input error message", "validation rules"],
        "23": ["language locale i18n translation", "language switch button"],
        "24": ["file upload download input", "file API endpoint"],
        "25": ["push notification alert", "notification API"],
        "26": ["concurrent user session multiple", "concurrent access"],
    }

    def generate_code_only(self, requirement: str, **kwargs) -> Dict[str, Any]:
        top_k = kwargs.get("top_k", 5)
        execution_id = kwargs.get("execution_id", 0)
        plan_path = kwargs.get("plan_path", None)
        scenario_serial = kwargs.get("scenario_serial", "")

        try:
            # 시나리오별 영어 쿼리로 다중 검색 후 컨텍스트 합산
            search_queries = self.SCENARIO_SEARCH_QUERIES.get(scenario_serial, [requirement])
            if not search_queries:
                search_queries = [requirement]

            contexts = []
            seen_chunks = set()
            per_query_chars = max(3000, 12000 // len(search_queries))

            for query in search_queries:
                ctx = self.rag_pipeline.retrieve_context(
                    requirement=query, top_k=top_k, max_chars=per_query_chars,
                    file_tree=getattr(self, "file_tree", None)
                )
                if ctx and ctx not in seen_chunks:
                    contexts.append(ctx)
                    seen_chunks.add(ctx)

            code_context = "\n\n".join(contexts)
            logger.info(f"[{execution_id}] Multi-query RAG: {len(search_queries)} queries, total context {len(code_context)} chars")

            test_cases = []
            logger.info(f"[{execution_id}] generate_code_only plan_path={plan_path}, exists={os.path.exists(plan_path) if plan_path else False}")
            if plan_path and os.path.exists(plan_path):
                try:
                    wb = openpyxl.load_workbook(plan_path)
                    ws = wb.active
                    for row in ws.iter_rows(min_row=6, values_only=True):
                        if row[4]:
                            test_cases.append({
                                "no": row[0],
                                "scenario_id": row[1],
                                "scenario_name": row[2],
                                "description": row[3],
                                "case_id": row[4],
                                "case_name": row[5],
                                "precondition": row[6],
                                "test_data": row[7],
                                "steps": row[8],
                                "expected_result": row[9],
                            })
                except Exception as e:
                    logger.error(f"[{execution_id}] Failed to load plan: {e}", exc_info=True)

            logger.info(f"[{execution_id}] STEP 2 — Generating test code from {len(test_cases)} cases (plan_path={plan_path})")
            test_plan_str = json.dumps(test_cases, ensure_ascii=False, indent=2)
            code_prediction = self.code_generator(test_plan_item=test_plan_str, code_context=code_context)
            test_code = self._parse_markdown_code(getattr(code_prediction, "generated_code", ""))

            # TypeScript → CommonJS 변환
            test_code = test_code.replace(
                "import { test, expect, Page } from '@playwright/test';",
                "const { test, expect } = require('@playwright/test');"
            )
            test_code = test_code.replace(
                "import { test, expect } from '@playwright/test';",
                "const { test, expect } = require('@playwright/test');"
            )
            test_code = re.sub(r'async \(\{ page \}: \{ page: Page \}\)', 'async ({ page })', test_code)
            test_code = re.sub(r'async \(page: Page\)', 'async (page)', test_code)

            timestamp = datetime.now().strftime("%Y%m%d")
            code_path = os.path.join(self.save_base_path, f"{timestamp}_{execution_id}.spec.js")
            with open(code_path, "w", encoding="utf-8") as f:
                f.write(test_code)
            logger.info(f"[{execution_id}] Saved test code: {code_path}")

            return {
                "status": "success",
                "saved_file": code_path,
                "saved_plan": plan_path,
                "test_code": test_code,
                "test_cases": test_cases,
            }

        except Exception as e:
            logger.error(f"[{execution_id}] generate_code_only failed: {e}", exc_info=True)
            return {"status": "error", "message": str(e)}

    # ── 내부 유틸리티 ──────────────────────────────────────────────────────────

    def _parse_test_plan(self, raw_plan: Any) -> List[Dict]:
        if isinstance(raw_plan, list):
            return raw_plan
        try:
            cleaned = str(raw_plan).strip()
            if cleaned.startswith("```"):
                lines = cleaned.split("\n")
                cleaned = "\n".join(lines[1:])
            if "```" in cleaned:
                cleaned = cleaned[:cleaned.index("```")]
            cleaned = cleaned.strip()
            start = cleaned.find("[")
            end = cleaned.rfind("]") + 1
            if start != -1 and end > start:
                cleaned = cleaned[start:end]
            return json.loads(cleaned.strip())
        except Exception as e:
            logger.error(f"_parse_test_plan failed: {e}, cleaned={str(cleaned)[:200]}")
            return []

    def _parse_markdown_code(self, raw_code: str) -> str:
        cleaned = raw_code.strip()
        if cleaned.startswith("```"):
            cleaned = "\n".join(cleaned.split("\n")[1:])
        if cleaned.endswith("```"):
            cleaned = "\n".join(cleaned.split("\n")[:-1])
        return cleaned.strip()

    def clear_index(self):
        self.rag_pipeline.clear()


# ── Excel 저장 ────────────────────────────────────────────────────────────────

def _save_plan_as_excel(test_cases: List[Dict], path: str) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SIT시나리오"

    thin = Side(style="thin", color="000000")
    border = Border(top=thin, bottom=thin, left=thin, right=thin)
    header_font = Font(name="Arial", bold=True, size=10)
    header_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    title_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
    title_font = Font(name="Arial", bold=True, size=12, color="FFFFFF")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    data_font = Font(name="Arial", size=10)

    ws.merge_cells("A1:N1")
    ws["A1"] = "SIT 시나리오"
    ws["A1"].font, ws["A1"].fill, ws["A1"].alignment, ws["A1"].border = title_font, title_fill, center, border
    ws.row_dimensions[1].height = 22

    for merge_range, value in [
        ("A4:A5", "No"), ("B4:B5", "테스트시나리오ID"), ("C4:C5", "테스트시나리오명"),
        ("D4:D5", "설명"), ("E4:J4", "테스트케이스"), ("K4:N4", "테스트 결과")
    ]:
        ws.merge_cells(merge_range)
        cell = ws[merge_range.split(":")[0]]
        cell.value, cell.font, cell.fill, cell.alignment, cell.border = value, header_font, header_fill, center, border

    for cell_ref, value in [
        ("E5", "테스트케이스ID"), ("F5", "테스트케이스명"), ("G5", "전제조건"),
        ("H5", "테스트 데이터"), ("I5", "실행단계"), ("J5", "예상결과"),
        ("K5", "테스트결과\n(화면)"), ("L5", "결과\n(정상/결함)"),
        ("M5", "테스트일자\n(YYYY-MM-DD)"), ("N5", "테스터")
    ]:
        cell = ws[cell_ref]
        cell.value, cell.font, cell.fill, cell.alignment, cell.border = value, header_font, header_fill, center, border

    row_idx = 6
    for tc in test_cases:
        ws.cell(row=row_idx, column=1, value=tc.get("no", ""))
        ws.cell(row=row_idx, column=2, value=tc.get("scenario_id", ""))
        ws.cell(row=row_idx, column=3, value=tc.get("scenario_name", ""))
        ws.cell(row=row_idx, column=4, value=tc.get("description", ""))

        for col, field in [(5, "case_id"), (6, "case_name"), (7, "precondition"),
                           (8, "test_data"), (9, "steps"), (10, "expected_result")]:
            val = tc.get(field, "")
            ws.cell(row=row_idx, column=col, value=str(val) if val else "")

        for col in range(1, 15):
            cell = ws.cell(row=row_idx, column=col)
            cell.font, cell.border, cell.alignment = data_font, border, (center if col in [1, 2, 3, 5, 7] else left)

        ws.row_dimensions[row_idx].height = 55
        row_idx += 1

    for col, width in zip("ABCDEFGHIJKLMN", [7, 18, 18, 50, 16, 20, 16, 25, 35, 25, 17, 11, 15, 11]):
        ws.column_dimensions[col].width = width

    ws.freeze_panes = "A6"
    wb.save(path)