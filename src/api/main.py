"""
FastAPI 서버

Spring Boot와 통신하는 AI 서버입니다.

엔드포인트:
  POST /api/generate-plan  — 레포 분석 → 테스트 계획서(Excel) 생성 → S3 → 콜백
  POST /api/execute-test   — 계획서 기반 테스트 코드 생성 → Playwright 실행 → 엑셀 결과 업데이트 → S3 → 콜백
  GET  /api/result/{execution_id} — 생성 결과 조회
  GET  /api/health — 서버 상태
"""

from fastapi import FastAPI, BackgroundTasks, Response, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
from typing import Optional, List, Dict
import asyncio
import logging
import tempfile
import shutil
import subprocess
import os
import re
import json
import time
import httpx
import boto3
import pickle
from datetime import datetime

from config.config import settings, validate_settings
from dotenv import load_dotenv
load_dotenv('/home/ec2-user/AI/.env')
from src.dspy_modules import configure_bedrock_dspy, RAGPlaywrightGenerator
from src.langchain_integration import CodeChunker
from langchain_core.documents import Document


STORE_PATH = "/home/ec2-user/AI/execution_store.pkl"

def _load_store() -> dict:
    if os.path.exists(STORE_PATH):
        try:
            with open(STORE_PATH, "rb") as f:
                return pickle.load(f)
        except Exception as e:
            return {}
    return {}

def _save_store():
    try:
        with open(STORE_PATH, "wb") as f:
            pickle.dump(execution_store, f)
    except Exception as e:
        logger.warning(f"Failed to save execution store: {e}")

execution_store: dict = _load_store()


logging.basicConfig(
    level=getattr(logging, settings.log_level.upper(), logging.INFO),
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s"
)
logger = logging.getLogger(__name__)

app = FastAPI(
    title="PROBE AI Server",
    description="AI 기반 Playwright 테스트 코드 자동 생성 서버",
    version="2.0.0"
)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


# ── 모델 ─────────────────────────────────────────────────────────────────────

class GeneratePlanRequest(BaseModel):
    model_config = {
        "populate_by_name": True,
        "alias_generator": lambda string: "".join(
            word.capitalize() if i else word
            for i, word in enumerate(string.split("_"))
        )
    }

    execution_id: int
    scenario_serial: str = ""
    scenario_attempt: str = ""
    target_branch: str = "main"
    repository_url: Optional[str] = None
    base_url: Optional[str] = None
    requirements: Optional[str] = None  # 단일 시나리오 번호 str
    server_url: Optional[str] = None
    auth_token: Optional[str] = None
    callback_url: str = ""
    group_name: Optional[str] = None  # 테스트명 (파일명에 사용)
    tester_name: Optional[str] = None  # 계정명 (엑셀 테스터 컬럼에 사용)

    @property
    def branch(self) -> str:
        return self.target_branch

    @property
    def scenario_id(self) -> str:
        return f"T{self.scenario_serial}{self.scenario_attempt}"


class ExecuteTestRequest(BaseModel):
    model_config = {
        "populate_by_name": True,
        "alias_generator": lambda string: "".join(
            word.capitalize() if i else word
            for i, word in enumerate(string.split("_"))
        )
    }

    execution_id: int
    callback_url: str = ""


class ScenarioDetail(BaseModel):
    scenarioName: Optional[str] = None
    description: Optional[str] = None
    testCaseId: Optional[str] = None
    testCaseName: Optional[str] = None
    precondition: Optional[str] = None
    testData: Optional[str] = None
    executionSteps: Optional[str] = None
    result: Optional[str] = None


class TestCaseResult(BaseModel):
    test_case_number: Optional[str] = None
    case_name: Optional[str] = None
    test_code_name: Optional[str] = None
    status: str  # "PASS" or "FAIL"
    duration_seconds: Optional[float] = None
    error_log: Optional[str] = None
    test_code: Optional[str] = None
    scenario_detail: Optional[ScenarioDetail] = None
    screenshot_s3_urls: Optional[List[str]] = None


class PlanCallbackPayload(BaseModel):
    execution_id: int
    plan_s3_url: Optional[str] = None


class TestCallbackPayload(BaseModel):
    execution_id: int
    status: str
    duration_seconds: Optional[float] = None
    plan_result_s3_url: Optional[str] = None
    test_spec_s3_url: Optional[str] = None
    results: Optional[List[TestCaseResult]] = None


# ── 공통 유틸 ─────────────────────────────────────────────────────────────────

EXCLUDE_DIRS = {
    ".git", "node_modules", "dist", "build", ".next",
    "__pycache__", "vendor", "target", "public", ".vite"
}

OUTPUT_DIR = "/home/ec2-user/AI/output_codes"

SCENARIO_NAME_MAP = {
    "01": "회원가입", "02": "로그인", "03": "비밀번호찾기", "04": "로그아웃",
    "05": "프로필수정", "06": "비밀번호변경", "07": "권한기반접근제어", "08": "세션만료",
    "09": "게시글작성", "10": "게시글수정삭제", "11": "댓글작성수정삭제", "12": "좋아요즐겨찾기",
    "13": "검색", "14": "필터정렬", "15": "반응형레이아웃", "16": "브라우저호환성",
    "17": "에러페이지동작", "18": "네트워크끊김", "19": "서버응답지연", "20": "API에러처리",
    "21": "AB테스트", "22": "입력값유효성검사", "23": "다국어지원", "24": "파일업로드다운로드",
    "25": "푸시알림", "26": "다중사용자동시접속",
}

def _get_scenario_filename(scenario_serial: str, file_type: str, date_str: str) -> str:
    """시나리오명 기반 파일명 생성"""
    name = SCENARIO_NAME_MAP.get(scenario_serial, f"시나리오{scenario_serial}")
    return f"{name}_{file_type}_{date_str}.xlsx"


def _clone_sync(repo_url: str, branch: str, token: Optional[str]) -> str:
    tmp_dir = tempfile.mkdtemp()
    url = repo_url
    if token:
        url = url.replace("https://", f"https://{token}@")
    try:
        subprocess.run(
            ["git", "clone", "-b", branch, "--depth", "1", url, tmp_dir],
            check=True, capture_output=True
        )
        return tmp_dir
    except subprocess.CalledProcessError as e:
        shutil.rmtree(tmp_dir)
        raise Exception(f"Git clone failed: {e.stderr.decode()}")


def _get_file_tree(path: str) -> str:
    tree = []
    for root, dirs, files in os.walk(path):
        dirs[:] = [d for d in dirs if d not in EXCLUDE_DIRS]
        level = root.replace(path, "").count(os.sep)
        tree.append(f"{'    ' * level}{os.path.basename(root)}/")
        for f in files:
            tree.append(f"{'    ' * (level + 1)}{f}")
    return "\n".join(tree)


def _load_documents(repo_path: str) -> List[Document]:
    documents = []
    target_extensions = (".js", ".ts", ".tsx", ".java", ".html", ".css")
    for root, dirs, files in os.walk(repo_path):
        dirs[:] = [d for d in dirs if d not in EXCLUDE_DIRS]
        for file in files:
            if file.endswith(target_extensions):
                file_path = os.path.join(root, file)
                if os.path.getsize(file_path) > 100 * 1024:
                    logger.warning(f"Skipping large file: {file_path}")
                    continue
                try:
                    with open(file_path, "r", encoding="utf-8") as f:
                        content = f.read()
                    rel_path = os.path.relpath(file_path, repo_path)
                    documents.append(Document(page_content=content, metadata={"source": rel_path}))
                except Exception as e:
                    logger.warning(f"Skipping {file_path}: {e}")
    return documents


async def _clone_and_chunk(repo_url: str, branch: str, auth_token: Optional[str], execution_id: int):
    repo_path = await asyncio.to_thread(_clone_sync, repo_url, branch, auth_token)

    logger.info(f"[{execution_id}] Scanning & loading files...")
    file_tree, documents = await asyncio.gather(
        asyncio.to_thread(_get_file_tree, repo_path),
        asyncio.to_thread(_load_documents, repo_path),
    )

    if not documents:
        shutil.rmtree(repo_path, ignore_errors=True)
        raise ValueError("분석할 소스 코드 파일(.js, .html 등)을 찾을 수 없습니다.")

    logger.info(f"[{execution_id}] Chunking documents...")
    chunker = CodeChunker(
        chunk_size=settings.rag_chunk_size,
        chunk_overlap=settings.rag_chunk_overlap
    )
    chunked_docs = await asyncio.to_thread(chunker.split_by_file_extension, documents)

    return repo_path, file_tree, chunked_docs


S3_BUCKET = "kiwi-test-artifacts-probe-2026"
S3_REGION = "us-east-1"


def _upload_to_s3(local_path: str, s3_key: str) -> Optional[str]:
    try:
        s3 = boto3.client("s3", region_name=S3_REGION)
        s3.upload_file(local_path, S3_BUCKET, s3_key)
        url = f"https://{S3_BUCKET}.s3.{S3_REGION}.amazonaws.com/{s3_key}"
        logger.info(f"Uploaded to S3: {url}")
        return url
    except Exception as e:
        logger.warning(f"S3 upload skipped: {e}")
        return None


def _create_white_png() -> bytes:
    """1x1 흰색 PNG 이미지 생성 (스크린샷 없을 때 대체용)"""
    import struct, zlib
    def make_png(w, h):
        def chunk(name, data):
            c = zlib.crc32(name + data) & 0xffffffff
            return struct.pack('>I', len(data)) + name + data + struct.pack('>I', c)
        raw = b'\x00' + b'\xff' * (w * 3)
        idat = zlib.compress(raw * h)
        return (b'\x89PNG\r\n\x1a\n'
                + chunk(b'IHDR', struct.pack('>IIBBBBB', w, h, 8, 2, 0, 0, 0))
                + chunk(b'IDAT', idat)
                + chunk(b'IEND', b''))
    return make_png(1, 1)


def _upload_bytes_to_s3(data: bytes, s3_key: str, content_type: str) -> Optional[str]:
    try:
        s3 = boto3.client("s3", region_name=S3_REGION)
        s3.put_object(Bucket=S3_BUCKET, Key=s3_key, Body=data, ContentType=content_type)
        url = f"https://{S3_BUCKET}.s3.{S3_REGION}.amazonaws.com/{s3_key}"
        logger.info(f"Uploaded bytes to S3: {url}")
        return url
    except Exception as e:
        logger.warning(f"S3 bytes upload skipped: {e}")
        return None


async def send_plan_callback(callback_url: str, payload: PlanCallbackPayload) -> None:
    try:
        async with httpx.AsyncClient(timeout=30) as client:
            resp = await client.post(
                callback_url,
                json=payload.model_dump(),
                headers={"Content-Type": "application/json"}
            )
            resp.raise_for_status()
            logger.info(f"Plan Callback OK → {callback_url} (execution_id={payload.execution_id})")
    except Exception as e:
        logger.error(f"Plan Callback FAILED → {callback_url} (execution_id={payload.execution_id}): {e}")


async def send_test_callback(callback_url: str, payload: TestCallbackPayload) -> None:
    try:
        async with httpx.AsyncClient(timeout=30) as client:
            resp = await client.post(
                callback_url,
                json=payload.model_dump(),
                headers={"Content-Type": "application/json"}
            )
            resp.raise_for_status()
            logger.info(f"Test Callback OK → {callback_url} (execution_id={payload.execution_id})")
    except Exception as e:
        logger.error(f"Test Callback FAILED → {callback_url} (execution_id={payload.execution_id}): {e}")


# ── PLAN 전용 함수 ────────────────────────────────────────────────────────────

def _run_plan_pipeline(requirement: str, execution_id: int, plan_index: int = 0, scenario_serial: str = "00", scenario_attempt: str = "01") -> dict:
    """STEP 1: 인덱싱 없이 시나리오명만으로 계획서 생성"""
    from src.dspy_modules.rag_generator import RAGPlaywrightGenerator as Gen
    generator = Gen(region=settings.aws_region)
    return generator.generate_plan_only(
        requirement=requirement,
        top_k=settings.rag_top_k,
        execution_id=execution_id,
        plan_index=plan_index,
        scenario_serial=scenario_serial,
        scenario_attempt=scenario_attempt,
    )


# ── TEST 전용 함수 ────────────────────────────────────────────────────────────

def _run_test_pipeline(chunked_docs, file_tree: str, requirement: str, execution_id: int, plan_path: Optional[str] = None, scenario_serial: str = "") -> dict:
    from src.dspy_modules.rag_generator import RAGPlaywrightGenerator as Gen
    generator = Gen(region=settings.aws_region)
    generator.index_documents(chunked_docs, file_tree=file_tree)
    return generator.generate_code_only(
        requirement=requirement,
        top_k=settings.rag_top_k,
        execution_id=execution_id,
        plan_path=plan_path,
        scenario_serial=scenario_serial,
    )


def _parse_test_cases_from_spec(spec_path: str) -> Dict[str, str]:
    """spec.js 파일에서 케이스ID별 테스트 코드 파싱"""
    result = {}
    if not spec_path or not os.path.exists(spec_path):
        return result
    try:
        with open(spec_path, "r", encoding="utf-8") as f:
            content = f.read()
        # test('T0201_01 - ...', async ({ page }) => { ... }); 패턴 파싱
        pattern = re.compile(
            r"test\((['\"])([^'\"]+)\1,\s*async\s*\([^)]*\)\s*=>\s*\{" ,
            re.MULTILINE
        )
        matches = list(pattern.finditer(content))
        for idx, match in enumerate(matches):
            case_title = match.group(2).strip()
            # 케이스 ID 추출 (T0201_01 형식)
            id_match = re.match(r"(T\d{4}_\d{2})", case_title)
            case_id = id_match.group(1) if id_match else case_title
            start = match.start()
            end = matches[idx + 1].start() if idx + 1 < len(matches) else len(content)
            result[case_id] = content[start:end].strip()
    except Exception as e:
        logger.warning(f"Failed to parse test cases from spec: {e}")
    return result


def _run_playwright(spec_path: str, execution_id: int) -> List[Dict]:
    result_dir = tempfile.mkdtemp()
    json_report_path = os.path.join(result_dir, "report.json")
    try:
        env = {
            **os.environ,
            "CI": "1",
            "PLAYWRIGHT_JSON_OUTPUT_NAME": json_report_path,
        }
        cmd = [
            "npx", "playwright", "test", spec_path,
            "--reporter=json",
            f"--output={result_dir}",
        ]
        proc = subprocess.run(
            cmd,
            timeout=600,
            cwd="/home/ec2-user/AI",
            env=env,
            capture_output=True,
        )
        logger.info(f"[{execution_id}] Playwright exit code: {proc.returncode}")

        if proc.stderr:
            stderr_preview = proc.stderr.decode(errors="ignore")[:300]
            if stderr_preview.strip():
                logger.warning(f"[{execution_id}] Playwright stderr: {stderr_preview}")

        results = []

        if os.path.exists(json_report_path) and os.path.getsize(json_report_path) > 0:
            try:
                with open(json_report_path) as jf:
                    report = json.load(jf)
                results = _parse_playwright_json(report)
                logger.info(f"[{execution_id}] Parsed {len(results)} results from JSON file")
            except Exception as e:
                logger.warning(f"[{execution_id}] JSON file parse failed: {e}")

        if not results and proc.stdout:
            stdout = proc.stdout.decode(errors="ignore").strip()
            try:
                results = _parse_playwright_json(json.loads(stdout))
                logger.info(f"[{execution_id}] Parsed {len(results)} results from stdout")
            except Exception:
                pass

        if not results:
            combined = ""
            if proc.stdout:
                combined += proc.stdout.decode(errors="ignore")
            if proc.stderr:
                combined += proc.stderr.decode(errors="ignore")
            results = _parse_playwright_text(combined, execution_id)

        logger.info(f"[{execution_id}] pw_results sample: {results[:2]}")
        return results

    except subprocess.TimeoutExpired:
        return [{"case_name": "Playwright Execution", "status": "FAIL",
                 "error_log": "Timeout", "screenshot_path": None}]
    except Exception as e:
        return [{"case_name": "Playwright Execution", "status": "FAIL",
                 "error_log": str(e), "screenshot_path": None}]
    finally:
        shutil.rmtree(result_dir, ignore_errors=True)


def _parse_playwright_json(report: dict) -> List[Dict]:
    results = []

    def walk_suite(suite):
        for spec in suite.get("specs", []):
            case_name = spec.get("title", "Unknown")
            ok = spec.get("ok", False)
            error_log = None
            duration_ms = 0
            for test in spec.get("tests", []):
                for result in test.get("results", []):
                    duration_ms += result.get("duration", 0)
                    if not ok and not error_log:
                        for err in result.get("errors", []):
                            msg = err.get("message", "")
                            error_log = msg[:300] if msg else str(err)
                            break
            results.append({
                "case_name": case_name,
                "status": "SUCCESS" if ok else "FAIL",
                "error_log": error_log,
                "duration_seconds": round(duration_ms / 1000, 1),
                "screenshot_path": None,
            })
        for sub in suite.get("suites", []):
            walk_suite(sub)

    for suite in report.get("suites", []):
        walk_suite(suite)

    return results or [{"case_name": "Playwright Execution", "status": "FAIL",
                        "error_log": "No results", "screenshot_path": None}]


def _parse_playwright_text(output: str, execution_id: int) -> List[Dict]:
    results = []
    for match in re.finditer(r'✓\s+\d+\s+(.+?)\s+\(\d+', output):
        results.append({"case_name": match.group(1).strip(), "status": "SUCCESS",
                        "error_log": None, "screenshot_path": None})
    for match in re.finditer(r'✘\s+\d+\s+(.+?)\s+\(\d+', output):
        results.append({"case_name": match.group(1).strip(), "status": "FAIL",
                        "error_log": "Test failed", "screenshot_path": None})
    if not results:
        passed = "passed" in output.lower()
        results.append({"case_name": "Playwright Execution",
                        "status": "SUCCESS" if passed else "FAIL",
                        "error_log": None if passed else output[-500:],
                        "screenshot_path": None})
    return results


def _update_excel_with_results(plan_path: str, pw_results: List[Dict], tester_name: str = "user") -> None:
    try:
        import openpyxl
        from openpyxl.drawing.image import Image as XLImage
        from openpyxl.styles import Font, Alignment, PatternFill
        from datetime import date

        normalized = []
        for r in pw_results:
            if isinstance(r, dict):
                normalized.append(r)
            elif isinstance(r, str):
                normalized.append({"case_name": r, "status": "FAIL", "error_log": None})
            else:
                normalized.append({"case_name": str(r), "status": "FAIL", "error_log": None})
        pw_results = normalized

        wb = openpyxl.load_workbook(plan_path)
        ws = wb.active
        today = date.today().strftime("%Y-%m-%d")
        center = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for i, res in enumerate(pw_results):
            row = 6 + i

            status = res.get("status", "FAIL")
            result_text = "정상" if status == "SUCCESS" else "결함"
            fill_color = "C6EFCE" if status == "SUCCESS" else "FFC7CE"
            font_color = "276221" if status == "SUCCESS" else "9C0006"

            cell_l = ws.cell(row=row, column=12, value=result_text)
            cell_l.font = Font(name="Arial", size=10, bold=True, color=font_color)
            cell_l.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
            cell_l.alignment = center

            ws.cell(row=row, column=13, value=today).alignment = center
            ws.cell(row=row, column=14, value=tester_name).alignment = center

            screenshot_path = os.path.join(
                "/home/ec2-user/AI/test-results",
                f"screenshot_{str(i+1).zfill(3)}.png"
            )
            if os.path.exists(screenshot_path):
                try:
                    img = XLImage(screenshot_path)
                    img.width = 160
                    img.height = 100
                    ws.row_dimensions[row].height = 80
                    ws.add_image(img, f"K{row}")
                except Exception as e:
                    logger.warning(f"Screenshot insert failed row {row}: {e}")
                    ws.cell(row=row, column=11, value="-").alignment = center
            else:
                ws.cell(row=row, column=11, value="-").alignment = center

        wb.save(plan_path)
        logger.info(f"Excel updated: {plan_path} ({len(pw_results)} rows)")
    except Exception as e:
        logger.error(f"Failed to update excel: {e}")
        raise


# ── 앱 이벤트 ────────────────────────────────────────────────────────────────

@app.on_event("startup")
async def startup_event():
    try:
        validate_settings()
        logger.info(f"Initializing DSPy with Bedrock in region: {settings.aws_region}")
        configure_bedrock_dspy(region=settings.aws_region)
        logger.info("DSPy initialization complete")
    except Exception as e:
        logger.error(f"Startup failed: {e}")
        raise


# ── 백그라운드 작업 ───────────────────────────────────────────────────────────

async def generate_plan_background(
    execution_id: int, repo_url: str, branch: str,
    requirements: str, callback_url: str,
    scenario_serial: str = "", scenario_attempt: str = "",
    auth_token: Optional[str] = None,
    server_url: Optional[str] = None,
    base_url: Optional[str] = None,
    group_name: Optional[str] = None,
    tester_name: Optional[str] = None,
):
    SCENARIO_MAP = {
        "01": "회원가입 시나리오 테스트 계획서 작성",
        "02": "로그인 시나리오 테스트 계획서 작성",
        "03": "비밀번호 찾기 시나리오 테스트 계획서 작성",
        "04": "로그아웃 시나리오 테스트 계획서 작성",
        "05": "프로필 수정 시나리오 테스트 계획서 작성",
        "06": "비밀번호 변경 시나리오 테스트 계획서 작성",
        "07": "권한 기반 접근 제어 시나리오 테스트 계획서 작성",
        "08": "세션 만료/토큰 만료 시나리오 테스트 계획서 작성",
        "09": "게시글 작성 시나리오 테스트 계획서 작성",
        "10": "게시글 수정/삭제 시나리오 테스트 계획서 작성",
        "11": "댓글 작성/수정/삭제 시나리오 테스트 계획서 작성",
        "12": "좋아요/즐겨찾기 시나리오 테스트 계획서 작성",
        "13": "검색 시나리오 테스트 계획서 작성",
        "14": "필터/정렬 시나리오 테스트 계획서 작성",
        "15": "반응형 레이아웃 시나리오 테스트 계획서 작성",
        "16": "브라우저 호환성 시나리오 테스트 계획서 작성",
        "17": "404/500 에러 페이지 동작 시나리오 테스트 계획서 작성",
        "18": "네트워크 끊김 상태 시나리오 테스트 계획서 작성",
        "19": "서버 응답 지연 시나리오 테스트 계획서 작성",
        "20": "API 에러 응답 처리 시나리오 테스트 계획서 작성",
        "21": "A/B 테스트 요소 확인 시나리오 테스트 계획서 작성",
        "22": "입력값 유효성 검사 시나리오 테스트 계획서 작성",
        "23": "다국어 지원 언어 변경 시나리오 테스트 계획서 작성",
        "24": "파일 업로드/다운로드 시나리오 테스트 계획서 작성",
        "25": "푸시 알림 시나리오 테스트 계획서 작성",
        "26": "다중 사용자 동시 접속 시나리오 테스트 계획서 작성",
    }

    started_at = time.time()
    try:


        # requirements가 없으면 scenario_serial로 대체, 둘 다 없으면 에러
        if not requirements:
            if not scenario_serial:
                raise ValueError("requirements와 scenario_serial 둘 다 없습니다. 요청을 확인하세요.")
            requirements = scenario_serial
            logger.info(f"[{execution_id}] [PLAN] requirements 미수신 → scenario_serial({scenario_serial})로 대체: {requirements}")

        # 번호 → 시나리오명 변환
        req_str = str(requirements).strip()
        resolved = SCENARIO_MAP.get(req_str, req_str)
        full_req = resolved
        if base_url and "base_url" not in full_req:
            full_req = f"base_url: {base_url}\n\n{full_req}"
        if server_url and "server_url" not in full_req:
            full_req = f"{full_req}\nserver_url: {server_url}"

        logger.info(f"[{execution_id}] [PLAN] requirements={requirements}, resolved={resolved}")
        logger.info(f"[{execution_id}] [PLAN] scenario_serial={scenario_serial}, scenario_attempt={scenario_attempt}")
        logger.info(f"[{execution_id}] [PLAN] Generating plan (no indexing)...")
        result = await asyncio.to_thread(
            _run_plan_pipeline, full_req,
            execution_id, 0, scenario_serial or "00", scenario_attempt or "01"
        )

        plan_path = None
        s3_url = None

        if not (isinstance(result, dict) and result.get("status") == "success"):
            logger.error(f"[{execution_id}] Plan failed: {result.get('message')}")
        else:
            plan_path = result.get("saved_plan", "")
            if plan_path and os.path.exists(plan_path):
                plan_date = datetime.now().strftime("%Y%m%d")
                plan_prefix = group_name if group_name else SCENARIO_NAME_MAP.get(scenario_serial or "00", f"시나리오{scenario_serial}")
                plan_filename = f"{plan_prefix}_계획서_{plan_date}.xlsx"
                s3_key = f"executions/{execution_id}/{plan_filename}"
                s3_url = await asyncio.to_thread(_upload_to_s3, plan_path, s3_key)

        execution_store[execution_id] = {
            "execution_id": execution_id,
            "status": "PLAN_COMPLETED",
            "plan_paths": [plan_path],
            "plan_s3_urls": [s3_url],
            "repo_url": repo_url,
            "branch": branch,
            "requirements": [full_req],
            "auth_token": auth_token,
            "scenario_serial": scenario_serial,
            "scenario_attempt": scenario_attempt,
            "server_url": server_url,
            "group_name": group_name or "",
            "tester_name": tester_name or "user",
            "tester_name": tester_name or "user",
        }
        _save_store()

        await send_plan_callback(callback_url, PlanCallbackPayload(
            execution_id=execution_id,
            plan_s3_url=s3_url,
        ))

    except Exception as e:
        logger.error(f"[{execution_id}] [PLAN] Job failed: {e}", exc_info=True)
        await send_plan_callback(callback_url, PlanCallbackPayload(
            execution_id=execution_id,
            plan_s3_url=None,
        ))
    finally:
        pass


async def execute_test_background(
    execution_id: int, callback_url: str,
):
    started_at = time.time()
    repo_path = None
    try:
        store = execution_store.get(execution_id)
        if not store:
            raise ValueError(f"execution_id {execution_id}에 해당하는 plan이 없습니다.")

        repo_url = store.get("repo_url", "")
        branch = store.get("branch", "main")
        requirements = store.get("requirements", [])
        server_url = store.get("server_url")
        auth_token = store.get("auth_token") or os.getenv("GITHUB_AUTH_TOKEN")
        scenario_serial = store.get("scenario_serial", "")
        scenario_attempt = store.get("scenario_attempt", "")
        scenario_id = f"T{scenario_serial}{scenario_attempt}"

        requirement = requirements[0] if requirements else ""

        if server_url and "server_url" not in requirement:
            requirement = f"{requirement}\nserver_url: {server_url}"

        logger.info(f"[{execution_id}] [TEST] scenario_id={scenario_id}, Cloning repository...")
        repo_path, file_tree, chunked_docs = await _clone_and_chunk(repo_url, branch, auth_token, execution_id)

        logger.info(f"[{execution_id}] [TEST] Generating test code...")

        plan_paths = store.get("plan_paths", [])
        plan_path_for_index = plan_paths[0] if plan_paths else None

        if not plan_path_for_index or not os.path.exists(plan_path_for_index):
            from datetime import datetime as dt
            today = dt.now().strftime("%Y%m%d")
            fallback = os.path.join(OUTPUT_DIR, f"{today}_{execution_id}_plan_1.xlsx")
            if os.path.exists(fallback):
                plan_path_for_index = fallback
                logger.info(f"[{execution_id}] Using fallback plan path: {fallback}")

        if not plan_path_for_index or not os.path.exists(plan_path_for_index):
            raise ValueError(f"계획서 파일을 찾을 수 없습니다. generate-plan을 먼저 실행하세요. (path={plan_path_for_index})")

        logger.info(f"[{execution_id}] plan_path_for_index={plan_path_for_index}, exists={os.path.exists(plan_path_for_index)}")
        result = await asyncio.to_thread(_run_test_pipeline, chunked_docs, file_tree, requirement, execution_id, plan_path_for_index, scenario_serial)

        if not (isinstance(result, dict) and result.get("status") == "success"):
            raise ValueError(result.get("message", "Test generation failed"))

        spec_path = result.get("saved_file", "")
        plan_path = result.get("saved_plan", "")

        logger.info(f"[{execution_id}] [TEST] Running Playwright...")
        pw_results = await asyncio.to_thread(_run_playwright, spec_path, execution_id)
        logger.info(f"[{execution_id}] [TEST] Playwright done: {len(pw_results)} results")

        if plan_path and os.path.exists(plan_path):
            await asyncio.to_thread(_update_excel_with_results, plan_path, pw_results, store.get("tester_name", "user"))

        report_s3_url = None
        if plan_path and os.path.exists(plan_path):
            result_date = datetime.now().strftime("%Y%m%d")
            result_serial = store.get("scenario_serial", "00")
            stored_group_name = store.get("group_name", "")
            result_prefix = stored_group_name if stored_group_name else SCENARIO_NAME_MAP.get(result_serial, f"시나리오{result_serial}")
            result_filename = f"{result_prefix}_결과보고서_{result_date}.xlsx"
            s3_key = f"executions/{execution_id}/{result_filename}"
            report_s3_url = await asyncio.to_thread(_upload_to_s3, plan_path, s3_key)

        spec_s3_url = None
        if spec_path and os.path.exists(spec_path):
            s3_key = f"executions/{execution_id}/test_1.spec.js"
            spec_s3_url = await asyncio.to_thread(_upload_to_s3, spec_path, s3_key)

        test_cases = result.get("test_cases", [])

        # spec 파일에서 케이스별 코드 파싱
        spec_code_map = _parse_test_cases_from_spec(spec_path)
        spec_filename = os.path.basename(spec_path) if spec_path else ""

        case_results = []
        for i, r in enumerate(pw_results):
            # case_name에서 test_case_number 파싱: "T0201_08 - 비밀번호..." → "T0201_08", "비밀번호..."
            raw_name = r.get("case_name", "")
            if " - " in raw_name:
                parsed_number, parsed_name = raw_name.split(" - ", 1)
                parsed_number = parsed_number.strip()
                parsed_name = parsed_name.strip()
            else:
                case_number = str(i + 1).zfill(2)
                parsed_number = f"{scenario_id}_{case_number}"
                parsed_name = raw_name

            screenshot_s3_urls = []
            screenshot_path = os.path.join(
                "/home/ec2-user/AI/test-results",
                f"screenshot_{str(i+1).zfill(3)}.png"
            )
            s3_key = f"executions/{execution_id}/screenshots/screenshot_{str(i+1).zfill(3)}.png"
            if os.path.exists(screenshot_path):
                with open(screenshot_path, "rb") as f:
                    screenshot_bytes = f.read()
            else:
                # 스크린샷 없으면 흰 이미지 전송
                screenshot_bytes = _create_white_png()
            url = await asyncio.to_thread(
                _upload_bytes_to_s3, screenshot_bytes, s3_key, "image/png"
            )
            if url:
                screenshot_s3_urls.append(url)

            tc = test_cases[i] if i < len(test_cases) else {}
            status_text = "정상" if r.get("status") == "SUCCESS" else "결함"
            scenario_detail = ScenarioDetail(
                scenarioName=tc.get("scenario_name"),
                description=tc.get("description"),
                testCaseId=tc.get("case_id", parsed_number),
                testCaseName=tc.get("case_name"),
                precondition=tc.get("precondition"),
                testData=tc.get("test_data"),
                executionSteps=tc.get("steps"),
                result=status_text,
            )

            is_pass = r.get("status") == "SUCCESS"
            case_test_code = spec_code_map.get(parsed_number)
            case_results.append(TestCaseResult(
                test_case_number=parsed_number,
                case_name=parsed_name,
                test_code_name=spec_filename,
                status="PASS" if is_pass else "FAIL",
                duration_seconds=r.get("duration_seconds"),
                error_log=r.get("error_log"),
                test_code=case_test_code,
                screenshot_s3_urls=screenshot_s3_urls,
                scenario_detail=scenario_detail,
            ))

        overall_status = "COMPLETED"
        if any(r.status == "FAIL" for r in case_results):
            logger.warning(f"[{execution_id}] Some tests failed.")

        execution_store[execution_id].update({
            "status": overall_status,
            "plan_s3_url": report_s3_url,
            "spec_s3_url": spec_s3_url,
            "results": [r.model_dump() for r in case_results],
        })
        _save_store()

        total_duration = round(time.time() - started_at, 1)

        # 전체 결과 한 번에 콜백 전송 (results를 JSON string으로)
        payload = TestCallbackPayload(
            execution_id=execution_id,
            status="COMPLETED",
            duration_seconds=total_duration,
            plan_result_s3_url=report_s3_url,
            test_spec_s3_url=spec_s3_url,
            results=case_results,
        )
        payload_dict = payload.model_dump()
        # test_code 길이만 로그에 찍기 (너무 길어서 전체 출력 불가)
        for r in payload_dict.get("results", []):
            if r.get("test_code"):
                r["test_code_len"] = len(r["test_code"])
                r["test_code"] = r["test_code"][:100] + "..."
        logger.info(f"[{execution_id}] Payload preview: {json.dumps(payload_dict, ensure_ascii=False)[:500]}")

        logger.info(f"[{execution_id}] Sending unified callback. Total cases: {len(case_results)}")
        await send_test_callback(callback_url, TestCallbackPayload(
            execution_id=execution_id,
            status="COMPLETED",
            duration_seconds=total_duration,
            plan_result_s3_url=report_s3_url,
            test_spec_s3_url=spec_s3_url,
            results=case_results,
        ))
        logger.info(f"[{execution_id}] Unified callback sent successfully.")

    except Exception as e:
        logger.error(f"[{execution_id}] [TEST] Job failed: {e}", exc_info=True)
        await send_test_callback(callback_url, TestCallbackPayload(
            execution_id=execution_id,
            status="FAILED",
            duration_seconds=round(time.time() - started_at, 1),
        ))
    finally:
        if repo_path and os.path.exists(repo_path):
            shutil.rmtree(repo_path)


# ── 엔드포인트 ───────────────────────────────────────────────────────────────

@app.post(
    "/api/generate-plan",
    status_code=202,
    summary="테스트 계획서 생성",
    description=(
        "GitHub 레포지토리를 분석하여 SIT 시나리오 형식의 테스트 계획서(Excel)를 생성합니다. "
        "executionId에는 본인이 사용할 값을 넣어주시면 됩니다. 동일한 executionId로 execute-test를 호출하여 계획서 기반 테스트 코드 생성 및 실행이 가능합니다. "
        "완료 시 callback_url로 결과를 전송합니다."
    ),
    responses={422: {"model": None}},
)
async def generate_plan(request: GeneratePlanRequest, background_tasks: BackgroundTasks) -> Response:
    background_tasks.add_task(
        generate_plan_background,
        request.execution_id,
        str(request.repository_url),
        request.target_branch,
        request.requirements,
        request.callback_url,
        request.scenario_serial,
        request.scenario_attempt,
        request.auth_token,
        request.server_url,
        str(request.base_url) if request.base_url else None,
        request.group_name,
        request.tester_name,
    )
    return Response(status_code=202)


@app.post(
    "/api/execute-test",
    status_code=202,
    summary="테스트 코드 생성 및 실행",
    description=(
        "generate-plan으로 생성된 계획서를 기반으로 Playwright 테스트 코드를 생성하고 실행합니다. "
        "실행 결과가 계획서 Excel에 반영되며, 완료 시 callback_url로 결과를 전송합니다. "
        "반드시 generate-plan을 먼저 실행해야 합니다."
    ),
    responses={422: {"model": None}},
)
async def execute_test(request: ExecuteTestRequest, background_tasks: BackgroundTasks) -> Response:
    if request.execution_id not in execution_store:
        fresh = _load_store()
        execution_store.update(fresh)

    if request.execution_id not in execution_store:
        raise HTTPException(
            status_code=404,
            detail=f"execution_id {request.execution_id}에 해당하는 plan이 없습니다. generate-plan을 먼저 실행하세요."
        )
    background_tasks.add_task(
        execute_test_background,
        request.execution_id,
        request.callback_url,
    )
    return Response(status_code=202)


@app.get(
    "/api/result/{execution_id}",
    summary="생성 결과 조회",
    description="테스트 계획서 S3 URL과 생성된 Playwright 테스트 코드를 조회합니다.",
)
async def get_result(execution_id: int):
    result = execution_store.get(execution_id)
    if not result:
        raise HTTPException(status_code=404, detail="Result not found. Job may still be running.")
    return result


@app.get(
    "/api/health",
    summary="서버 상태 확인",
    responses={200: {"content": {"application/json": {"example": {"status": "healthy"}}}}},
)
async def health_check():
    return {
        "status": "healthy",
        "region": settings.aws_region,
        "model": settings.bedrock_model,
        "environment": settings.environment,
    }


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(
        "main:app",
        host=settings.fastapi_host,
        port=settings.fastapi_port,
        reload=(settings.environment == "development"),
    )